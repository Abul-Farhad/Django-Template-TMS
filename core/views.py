from django.shortcuts import render, redirect
from django.contrib import messages
from django.views import View
from django.core import serializers
from django.http import JsonResponse, StreamingHttpResponse
from openpyxl import Workbook
from accounts.models import CustomUser
import csv
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
# def dashboard_view(request):
#     if request.user.is_authenticated:
#         return render(request, 'dashboard.html')
#     else:
#         messages.warning(request, 'You must be logged in to access this page.')
#         return redirect('login')  # Replace 'login' with your login URL name
    
class DashboardView(View):
    def get(self, request):
        if request.user.is_authenticated:
            users_count = CustomUser.objects.count()  # Count the number of users
            return render(request, 'dashboard.html', {'users_count': users_count})
        else:
            messages.warning(request, 'You must be logged in to access this page.')
            return redirect('login')
        
class ListUsersView(View):
    def get(self, request):
        search_query = request.GET.get('q', '')  # Get the search query from the request
        if search_query:
            # Filter users by name (case-insensitive)
            users = CustomUser.objects.filter(name__icontains=search_query).order_by('id')
        else:
            # List all users if no search query is provided
            users = CustomUser.objects.all().order_by('id')
        return render(request, 'users.html', {'users': users})
    



class Echo:
    """A helper class to stream CSV rows as they're written."""
    def write(self, value):
        return value


class ExportUserCSVView(View):
    def get(self, request):
        if not request.user.is_authenticated or not request.user.is_staff:
            return HttpResponse(status=403)

        # Define the streaming response

        response = StreamingHttpResponse(
            self.stream_users(request),
            content_type='text/csv'
        )
        response['Content-Disposition'] = 'attachment; filename="users.csv"'
        return response

    def stream_users(self, request):

        name = request.GET.get('name', '')
        email = request.GET.get('email', '')
        role = request.GET.get('role', '')

        # Handle ordering
        order_column_index = request.GET.get('order[0][column]', 0)
        order_dir = request.GET.get('order[0][dir]', 'asc')
        columns = ['name', 'email', 'is_staff', 'is_superuser', 'role__name']
        # print("Columns fields: ", columns)
        order_column = columns[int(order_column_index)]
        if order_dir == 'desc':
            order_column = '-' + order_column
        
        # print("Order Column: ", order_column)
        # print("Order Column Index: ", order_column_index)

        # Queryset
        users = CustomUser.objects.select_related('role')

        if name:
            users = users.filter(name__icontains=name)
        if email:
            users = users.filter(email__icontains=email)
        if role:
            users = users.filter(role__name__icontains=role)
        
        users = users.order_by(order_column)

        """Generator that yields CSV lines."""
        pseudo_buffer = Echo()
        writer = csv.writer(pseudo_buffer)

        # Write header first
        yield writer.writerow(['Name', 'Email', 'Is Staff', 'Is Superuser', 'Role'])

        # Stream results using iterator
        queryset = users.iterator(chunk_size=1000)
        for user in queryset:
            yield writer.writerow([
                user.name,
                user.email,
                'Yes' if user.is_staff else 'No',
                'Yes' if user.is_superuser else 'No',
                user.role.name if user.role else '—'
            ])





class ExportUserEXELView(View):
    def get(self, request):
        if not request.user.is_authenticated or not request.user.is_staff:
            return HttpResponse(status=403)

        # Create a new Excel workbook and active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"

        # Define styles
        header_font = Font(size=12, bold=True)
        odd_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # light gray
        even_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # white

        # Add header row with styles
        headers = ['Name', 'Email', 'Is Staff', 'Is Superuser', 'Role']
        ws.append(headers)
        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font

        # Populate user data
        users = CustomUser.objects.select_related('role').all()
        for idx, user in enumerate(users, start=2):
            row = [
                user.name,
                user.email,
                'Yes' if user.is_staff else 'No',
                'Yes' if user.is_superuser else 'No',
                user.role.name if user.role else '—'
            ]
            ws.append(row)

            # Apply alternating row fill
            fill = odd_fill if idx % 2 else even_fill
            for col_num in range(1, len(row) + 1):
                ws.cell(row=idx, column=col_num).fill = fill

        # Generate and return the Excel file as response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=users.xlsx'
        wb.save(response)
        return response


        